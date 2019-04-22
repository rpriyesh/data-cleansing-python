# -*- coding: utf-8 -*-

# Author:   Priyesh Rajamohanan
# Date:     27 November 2016
# Desc:     Excel file operations using OpenPyXL

from openpyxl import load_workbook
from openpyxl import Workbook
from customer_order import CustomerOrder
from openpyxl.styles import NamedStyle, Font
import os


class FileReader:
    def __init__(self, file_name):
        self.file_name = file_name

    def read_file(self):
        # dictionary to store order data
        order_details = {}
        order_wb = load_workbook(filename=self.file_name)
        order_ws = order_wb.active
        counter = -2
        for row in order_ws.iter_rows():
            counter += 1
            # skipping first 2 rows
            if counter < 1:
                continue
            order = CustomerOrder(row[0].value, row[1].value, row[2].value, row[3].value,
                                  row[4].value, row[5].value, row[6].value)
            order_details[counter] = order
        return order_details


class FileWriter:
    def __init__(self, file_name):
        self.file_name = file_name

    def delete_file_if_exists(self):
        if os.path.exists(self.file_name):
            try:
                os.remove(self.file_name)
                return True
            except WindowsError:
                return False
        else:
            return True

    def write_file(self, order_details):
        if self.delete_file_if_exists():
            order_wb = Workbook()
            order_ws = order_wb.active
            # set sheet name
            order_ws.title = "Order Data (2014-2016)"

            # defining column styling
            header_style = NamedStyle(name="headerStyle")
            header_style.font = Font(bold=True)

            date_style = NamedStyle(name="dateStyle")
            date_style.number_format = 'MM/DD/YYYY'

            amount_style = NamedStyle(name="amountStyle")
            amount_style.number_format = u'_-"€"* #,##0.00_-;-"€"* #,##0.00_-;_-"€"* "-"??_-;_-@_-'

            order_ws.append(["Contact Name", "Company", "Address 1", "County", "Country Code",
                             "Phone", "Amount", "Order Date"])

            # applying styles to header columns
            order_ws["A1"].style = header_style
            order_ws.column_dimensions["A"].width = 25

            order_ws["B1"].style = header_style
            order_ws.column_dimensions["B"].width = 25

            order_ws["C1"].style = header_style
            order_ws.column_dimensions["C"].width = 40

            order_ws["D1"].style = header_style
            order_ws.column_dimensions["D"].width = 15

            order_ws["E1"].style = header_style
            order_ws.column_dimensions["E"].width = 12

            order_ws["F1"].style = header_style
            order_ws.column_dimensions["F"].width = 10

            order_ws["G1"].style = header_style
            order_ws.column_dimensions["G"].width = 14

            order_ws["H1"].style = header_style
            order_ws.column_dimensions["H"].width = 12

            max_count = len(order_details)
            row_index = 1
            for index in range(1, max_count, 1):
                order = order_details[index]
                # print vars(order)
                # adding only valid rows to sheet
                if order.is_valid:
                    row = (order.contact_name, order.company, order.address1, order.county,
                           order.country_code, order.phone, order.amount, order.order_date)
                    order_ws.append(row)

                    # applying styles to Order Date and Amount columns
                    order_ws["H" + str(row_index + 1)].style = date_style
                    order_ws["G" + str(row_index + 1)].style = amount_style
                    row_index += 1
            order_wb.save(self.file_name)
        else:
            print ("Unable to delete file. Please close the file if it is already open.")
